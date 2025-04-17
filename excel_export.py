import pandas as pd
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
import database as db
from io import BytesIO
import streamlit as st

def create_report_excel(template_id, assigned_report_id, submission_data):
    """
    Create an Excel file for a report submission with multiple sheets based on template configuration.
    
    Args:
        template_id: ID of the report template
        assigned_report_id: ID of the assigned report
        submission_data: JSON string of submitted data
        
    Returns:
        BytesIO object containing the Excel file
    """
    # Get report template details
    template = db.get_report_template(template_id)
    if not template:
        raise ValueError("Không tìm thấy mẫu báo cáo")
    
    # Get the sheet structure or create a default one
    sheet_structure = db.get_report_template_sheet_structure(template_id)
    if sheet_structure:
        sheets = json.loads(sheet_structure)
    else:
        # Default structure with all fields in one sheet
        fields = json.loads(template['fields'])
        sheets = {"Sheet1": fields}
    
    # Parse submission data
    data = json.loads(submission_data)
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066b2", end_color="0066b2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # For each sheet in the structure
    for sheet_name, sheet_fields in sheets.items():
        # Create a new sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        ws.append(["Trường", "Giá trị"])
        
        # Style headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for field in sheet_fields:
            value = data.get(field, "")
            ws.append([field, value])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50
    
    # Save to BytesIO
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    return excel_bytes

def save_excel_to_sharepoint(excel_bytes, template_name, organization_name, assigned_report_id):
    """
    Save Excel file to SharePoint (mock implementation)
    
    Args:
        excel_bytes: BytesIO object containing the Excel file
        template_name: Name of the report template
        organization_name: Name of the organization
        assigned_report_id: ID of the assigned report
        
    Returns:
        URL to the saved file in SharePoint
    """
    # Get SharePoint settings
    sharepoint_settings = load_sharepoint_settings()
    
    # In a real implementation, this would upload the file to SharePoint
    # For now, we'll just return a mock URL
    sharepoint_url = sharepoint_settings.get("sharepoint_url", "https://vinatex.sharepoint.com")
    document_library = sharepoint_settings.get("document_library", "Documents/Reports")
    use_org_folders = sharepoint_settings.get("use_org_folders", True)
    
    # Create the folder structure
    folder_path = document_library
    if use_org_folders:
        folder_path = f"{folder_path}/{organization_name}"
    
    # Create filename
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{template_name}_{timestamp}.xlsx"
    
    # Mock URL
    file_url = f"{sharepoint_url}/{folder_path}/{filename}"
    
    # In a real implementation, we would:
    # 1. Connect to SharePoint using the credentials
    # 2. Create the folder structure if it doesn't exist
    # 3. Upload the file
    # 4. Return the URL to the uploaded file
    
    # For now, save URL to the database
    db.execute_query(
        "UPDATE report_submissions SET sharepoint_url = %s WHERE assigned_report_id = %s",
        (file_url, assigned_report_id),
        fetch=False
    )
    
    return file_url

def load_sharepoint_settings():
    """Load SharePoint settings from the database"""
    settings = db.get_settings("sharepoint")
    
    if settings:
        try:
            return json.loads(settings)
        except:
            pass
    
    # Default settings
    return {
        "sharepoint_url": "https://vinatex.sharepoint.com/sites/reports",
        "document_library": "Documents/Reports",
        "sharepoint_username": "admin@vinatex.com.vn",
        "sharepoint_password": "",
        "use_org_folders": True
    }

def export_report_to_excel(assigned_report_id):
    """
    Export a report to Excel and save to SharePoint
    
    Args:
        assigned_report_id: ID of the assigned report
        
    Returns:
        URL to the saved file in SharePoint
    """
    # Get report submission
    submission = db.get_report_submission(assigned_report_id)
    if not submission:
        st.error("Không tìm thấy báo cáo")
        return None
    
    # Get assigned report details
    assigned_report = db.execute_query(
        """
        SELECT ar.template_id, rt.name as template_name, o.name as organization_name
        FROM assigned_reports ar
        JOIN report_templates rt ON ar.template_id = rt.id
        JOIN organizations o ON ar.organization_id = o.id
        WHERE ar.id = %s
        """,
        (assigned_report_id,)
    )
    
    if assigned_report is None or assigned_report.empty:
        st.error("Không tìm thấy thông tin báo cáo")
        return None
    
    template_id = assigned_report.iloc[0]['template_id']
    template_name = assigned_report.iloc[0]['template_name']
    organization_name = assigned_report.iloc[0]['organization_name']
    
    # Create Excel file
    try:
        excel_bytes = create_report_excel(template_id, assigned_report_id, submission['data'])
        
        # Save to SharePoint
        file_url = save_excel_to_sharepoint(excel_bytes, template_name, organization_name, assigned_report_id)
        
        return file_url
    except Exception as e:
        st.error(f"Lỗi khi tạo file Excel: {str(e)}")
        return None

def download_report_excel(assigned_report_id):
    """
    Generate Excel file for download
    
    Args:
        assigned_report_id: ID of the assigned report
        
    Returns:
        BytesIO object containing the Excel file and filename
    """
    # Get report submission
    submission = db.get_report_submission(assigned_report_id)
    if not submission:
        st.error("Không tìm thấy báo cáo")
        return None, None
    
    # Get assigned report details
    assigned_report = db.execute_query(
        """
        SELECT ar.template_id, rt.name as template_name, o.name as organization_name
        FROM assigned_reports ar
        JOIN report_templates rt ON ar.template_id = rt.id
        JOIN organizations o ON ar.organization_id = o.id
        WHERE ar.id = %s
        """,
        (assigned_report_id,)
    )
    
    if assigned_report is None or assigned_report.empty:
        st.error("Không tìm thấy thông tin báo cáo")
        return None, None
    
    template_id = assigned_report.iloc[0]['template_id']
    template_name = assigned_report.iloc[0]['template_name']
    organization_name = assigned_report.iloc[0]['organization_name']
    
    # Create Excel file
    try:
        excel_bytes = create_report_excel(template_id, assigned_report_id, submission['data'])
        
        # Create filename
        timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{template_name}_{organization_name}_{timestamp}.xlsx"
        
        return excel_bytes, filename
    except Exception as e:
        st.error(f"Lỗi khi tạo file Excel: {str(e)}")
        return None, None