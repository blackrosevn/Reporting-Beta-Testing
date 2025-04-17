import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import json

def create_report_template(fields):
    """Create an Excel template with the specified fields."""
    # Create a new workbook and select the active worksheet
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Report Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add header row
    for col, field in enumerate(fields, 1):
        cell = worksheet.cell(row=1, column=col, value=field)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add one empty row for data entry
    for col in range(1, len(fields) + 1):
        cell = worksheet.cell(row=2, column=col, value="")
        cell.border = thin_border
    
    # Adjust column widths
    for col in range(1, len(fields) + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()

def parse_excel_report(uploaded_file, expected_fields):
    """Parse an Excel report and extract the field values."""
    # Read the Excel file
    df = pd.read_excel(uploaded_file)
    
    # Validate that all expected fields are present as column headers
    for field in expected_fields:
        if field not in df.columns:
            raise ValueError(f"Missing expected field: {field}")
    
    # Extract the first row of data for each field
    data = {}
    if len(df) > 0:
        for field in expected_fields:
            # Convert to string to ensure compatibility with JSON
            data[field] = str(df[field].iloc[0])
    else:
        raise ValueError("Excel file contains no data rows")
    
    return data

def create_excel_from_report(report, submission):
    """Create an Excel file from a report submission."""
    # Create a new workbook and select the active worksheet
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = report['report_name']
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=12)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add report title
    worksheet.merge_cells('A1:C1')
    title_cell = worksheet.cell(row=1, column=1, value=report['report_name'])
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add report details
    worksheet.cell(row=3, column=1, value="Due Date:").font = Font(bold=True)
    worksheet.cell(row=3, column=2, value=str(report['due_date']))
    
    worksheet.cell(row=4, column=1, value="Status:").font = Font(bold=True)
    worksheet.cell(row=4, column=2, value=report['status'])
    
    if submission:
        worksheet.cell(row=5, column=1, value="Submitted at:").font = Font(bold=True)
        worksheet.cell(row=5, column=2, value=str(submission['submitted_at']))
    
    # Add report data
    row = 7
    worksheet.cell(row=row, column=1, value="Report Data").font = subtitle_font
    row += 2
    
    if submission:
        try:
            data = json.loads(submission['data'])
            
            # Add header row
            col = 1
            for field in data.keys():
                cell = worksheet.cell(row=row, column=col, value=field)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                col += 1
            
            # Add data row
            row += 1
            col = 1
            for value in data.values():
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.border = thin_border
                col += 1
        except:
            worksheet.cell(row=row, column=1, value="Error parsing submission data")
    
    # Adjust column widths
    for col in range(1, 10):  # Adjust up to 10 columns
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()

def create_status_report(reports_df):
    """Create an Excel status report from the reports dataframe."""
    # Create a new workbook and select the active worksheet
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Report Status"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066B2", end_color="0066B2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    completed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    overdue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    worksheet.merge_cells('A1:E1')
    title_cell = worksheet.cell(row=1, column=1, value="Vinatex Report Status")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Add generation date
    import datetime
    worksheet.merge_cells('A2:E2')
    date_cell = worksheet.cell(row=2, column=1, value=f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    date_cell.alignment = Alignment(horizontal="center")
    
    # Add header row
    headers = ["Report Name", "Organization", "Due Date", "Status", "ID"]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add data rows
    for i, (_, row) in enumerate(reports_df.iterrows(), 5):
        worksheet.cell(row=i, column=1, value=row['report_name']).border = thin_border
        worksheet.cell(row=i, column=2, value=row['organization']).border = thin_border
        worksheet.cell(row=i, column=3, value=str(row['due_date'])).border = thin_border
        
        status_cell = worksheet.cell(row=i, column=4, value=row['status'])
        status_cell.border = thin_border
        
        # Color-code the status
        if row['status'] == 'completed':
            status_cell.fill = completed_fill
        elif row['status'] == 'pending':
            status_cell.fill = pending_fill
        elif row['status'] == 'overdue':
            status_cell.fill = overdue_fill
        
        worksheet.cell(row=i, column=5, value=row['id']).border = thin_border
    
    # Hide the ID column (used for reference)
    worksheet.column_dimensions['E'].hidden = True
    
    # Adjust column widths
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    
    # Add filters
    worksheet.auto_filter.ref = f"A4:D{len(reports_df) + 4}"
    
    # Save the workbook to the BytesIO object
    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()
