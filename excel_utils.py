import os
import json
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import database as db
import streamlit as st

def create_excel_from_template(template_id, data):
    """
    Create an Excel file with multiple sheets based on the template's sheet structure.
    
    Args:
        template_id: The ID of the report template
        data: JSON data of field values
        
    Returns:
        BytesIO object containing the Excel file
    """
    # Get the template details
    template = db.get_report_template(template_id)
    if not template:
        raise ValueError("Mẫu báo cáo không tồn tại")
    
    # Get sheet structure or create default
    sheet_structure = db.get_report_template_sheet_structure(template_id)
    if sheet_structure:
        sheets = json.loads(sheet_structure)
    else:
        # Default structure with all fields in one sheet
        fields = json.loads(template['fields'])
        sheets = {"Sheet1": fields}
    
    # Parse the submitted data
    if isinstance(data, str):
        field_values = json.loads(data)
    else:
        field_values = data
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color="0066b2", end_color="0066b2", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Create each sheet according to the structure
    for sheet_name, sheet_fields in sheets.items():
        # Create new sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add header row with styling
        for col_idx, field in enumerate(sheet_fields, 1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add data
        for col_idx, field in enumerate(sheet_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field_values.get(field, ""))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
        
        # Auto-adjust column widths
        for col_idx, _ in enumerate(sheet_fields, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20
    
    # Create a BytesIO object to store the workbook
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def parse_excel_submission(uploaded_file, template_id):
    """
    Parse an uploaded Excel file that was created from a template.
    
    Args:
        uploaded_file: The uploaded Excel file
        template_id: The ID of the report template
        
    Returns:
        Dictionary of field values
    """
    # Get the template details
    template = db.get_report_template(template_id)
    if not template:
        raise ValueError("Mẫu báo cáo không tồn tại")
    
    # Get sheet structure
    sheet_structure = db.get_report_template_sheet_structure(template_id)
    if sheet_structure:
        expected_sheets = json.loads(sheet_structure)
    else:
        fields = json.loads(template['fields'])
        expected_sheets = {"Sheet1": fields}
    
    # Load the workbook
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    
    # Extract data from each sheet
    field_values = {}
    
    for sheet_name, expected_fields in expected_sheets.items():
        if sheet_name not in wb.sheetnames:
            st.warning(f"Thiếu sheet {sheet_name} trong file Excel. Dữ liệu có thể không đầy đủ.")
            continue
        
        ws = wb[sheet_name]
        
        # Get header row (first row)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        # Check if all expected fields are present
        missing_fields = set(expected_fields) - set(headers)
        if missing_fields:
            st.warning(f"Thiếu các trường sau trong sheet {sheet_name}: {', '.join(missing_fields)}")
        
        # Get the data from row 2
        if len(ws.rows) < 2:
            st.warning(f"Không có dữ liệu trong sheet {sheet_name}")
            continue
        
        # Map field values
        for col_idx, header in enumerate(headers, 1):
            if header in expected_fields:
                value = ws.cell(row=2, column=col_idx).value
                field_values[header] = value if value is not None else ""
    
    return field_values

def create_excel_template(template_id):
    """
    Create an empty Excel template based on a report template.
    
    Args:
        template_id: The ID of the report template
        
    Returns:
        BytesIO object containing the Excel template
    """
    # Get the template details
    template = db.get_report_template(template_id)
    if not template:
        raise ValueError("Mẫu báo cáo không tồn tại")
    
    # Get sheet structure
    sheet_structure = db.get_report_template_sheet_structure(template_id)
    if sheet_structure:
        sheets = json.loads(sheet_structure)
    else:
        # Default structure with all fields in one sheet
        fields = json.loads(template['fields'])
        sheets = {"Sheet1": fields}
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color="0066b2", end_color="0066b2", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Create each sheet according to the structure
    for sheet_name, sheet_fields in sheets.items():
        # Create new sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # Add report name and instructions
        ws.merge_cells('A1:D1')
        title_cell = ws.cell(row=1, column=1, value=f"BÁO CÁO: {template['name']}")
        title_cell.font = Font(name='Arial', bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:D2')
        instructions = ws.cell(row=2, column=1, value="Hướng dẫn: Điền thông tin vào hàng dưới các tiêu đề. Không thay đổi cấu trúc file.")
        instructions.font = Font(name='Arial', italic=True, size=11)
        
        # Add header row with styling (row 4)
        for col_idx, field in enumerate(sheet_fields, 1):
            cell = ws.cell(row=4, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add empty data row
        for col_idx, _ in enumerate(sheet_fields, 1):
            cell = ws.cell(row=5, column=col_idx, value="")
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
        
        # Auto-adjust column widths
        for col_idx, _ in enumerate(sheet_fields, 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20
    
    # Create a BytesIO object to store the workbook
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def save_to_sharepoint(excel_file, template_name, org_name):
    """
    Simulate saving to SharePoint - in a real application this would use
    SharePoint API to save the file to the configured location.
    
    Args:
        excel_file: The Excel file to save (BytesIO)
        template_name: The name of the report template
        org_name: The name of the organization
        
    Returns:
        String URL where the file would be saved
    """
    # In a real implementation, this would use the SharePoint API
    # For now, return a placeholder URL
    
    # Get SharePoint settings
    sharepoint_settings = db.get_settings("sharepoint")
    
    if sharepoint_settings:
        settings = json.loads(sharepoint_settings)
        base_url = settings.get("sharepoint_url", "https://vinatex.sharepoint.com/sites/reports")
        document_library = settings.get("document_library", "Documents/Reports")
        use_org_folders = settings.get("use_org_folders", True)
    else:
        base_url = "https://vinatex.sharepoint.com/sites/reports"
        document_library = "Documents/Reports"
        use_org_folders = True
    
    # Create a sanitized filename
    safe_template_name = template_name.replace(" ", "_")
    safe_org_name = org_name.replace(" ", "_")
    
    # Generate timestamp
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Construct the path
    if use_org_folders:
        path = f"{document_library}/{safe_org_name}/{safe_template_name}_{timestamp}.xlsx"
    else:
        path = f"{document_library}/{safe_template_name}_{safe_org_name}_{timestamp}.xlsx"
    
    # Construct the full URL
    sharepoint_url = f"{base_url}/{path}"
    
    return sharepoint_url